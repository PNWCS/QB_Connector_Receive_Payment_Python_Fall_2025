from __future__ import annotations
from .models import CustomerReceivePaymentTerms
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
from typing import List


def read_CustomerReceivePaymentTerms_from_excel(
    file_path: Path,
) -> List[CustomerReceivePaymentTerms]:
    workbook_path = Path(file_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    workbook = load_workbook(filename=workbook_path, data_only=True)
    try:
        sheet = workbook["account credit vendor"]
    except KeyError as exc:
        workbook.close()
        raise ValueError(
            "Worksheet 'account credit vendor' not found in workbook"
        ) from exc
    rows = sheet.iter_rows(values_only=True)
    headers_row = next(rows, None)  # First row should contain column headers
    if headers_row is None:  # Empty sheet edge case
        workbook.close()
        return []

    # Build a mapping from header name to its column index
    headers = [
        str(header).strip() if header is not None else "" for header in headers_row
    ]
    header_index = {header: idx for idx, header in enumerate(headers)}

    def _value(row, column_name: str):  # Helper to safely access a column
        idx = header_index.get(column_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    terms: List[CustomerReceivePaymentTerms] = []  # Accumulator for valid terms
    try:
        for row in rows:  # Iterate over each data row
            record_id = _value(
                row, "Child ID"
            )  # Expected ID column (e.g., number of days)

            name = _value(row, "Customer")  # Expected Name column
            if name is None:
                continue  # Skip rows without a name

            if record_id in (None, ""):
                continue  # Skip rows without an ID

            if not record_id:
                continue  # Skip empty/invalid IDs
            date = _value(row, "Bank Date")
            if date is None or (isinstance(date, str) and not date.strip()):
                date = None
            elif isinstance(date, datetime):
                date = date.date().isoformat()
            # elif isinstance(date, date_cls):
            #     date = date.isoformat()
            else:
                # assume string in YYYY-MM-DD; adjust if needed
                date = datetime.strptime(str(date), "%Y-%m-%d").date().isoformat()
            amount = _value(row, "Check Amount")
            if not amount:
                continue  # Skip rows with invalid amount
            invoice_number = _value(row, "Invoice Number")
            if not invoice_number:
                continue  # Skip rows with invalid invoice number
            # Construct the domain object tagged as sourced from Excel

            terms.append(
                CustomerReceivePaymentTerms(
                    child_id=int(record_id),
                    customer=name,
                    date=date,
                    amount=float(amount),
                    invoice_number=invoice_number,
                    source="excel",
                )
            )
    finally:
        workbook.close()  # Always close the workbook handle

    return terms
